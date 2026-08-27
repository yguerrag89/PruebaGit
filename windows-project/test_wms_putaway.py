from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from core.live import LiveUnload
from core.optimizer import Settings
from core.parser import CodeRecord
from core.wms_putaway import (
    WMS_HEADERS,
    build_putaway_rows,
    export_official_putaway_xlsx,
)


APP_DIR = Path(__file__).resolve().parent
TEMPLATE = APP_DIR / "assets" / "templates" / "Plantilla_oficial_WMS_PutawayCrossDockImport.xlsx"


def accepted(scan, code, pallet="I01-01", position="I01"):
    return {
        "Escaneo": scan,
        "Código": code,
        "Posición": position,
        "Tarima": pallet,
        "Estado": "OK",
        "Caja individual": True,
    }


records = {
    "MJ260510161": CodeRecord("MJ260510161", 19, 1.9, 0.1),
    "MJ260561713": CodeRecord("MJ260561713", 6, 0.6, 0.1),
}
events = [
    accepted("MJ260510161U013", "MJ260510161"),
    accepted("MJ260561713U002", "MJ260561713"),
]


# 1) Caso dorado de las dos cajas pendientes.
result = build_putaway_rows(
    events,
    records,
    "PAS3902608080RT",
    default_location="1-1-01",
    received=2,
    expected=949,
    allow_partial=True,
)
assert result.ready, result.errors
assert result.rows == [
    {
        WMS_HEADERS[0]: "PAS3902608080RT",
        WMS_HEADERS[1]: "MJ260510161U013",
        WMS_HEADERS[2]: 1,
        WMS_HEADERS[3]: "1-1-01",
    },
    {
        WMS_HEADERS[0]: "PAS3902608080RT",
        WMS_HEADERS[1]: "MJ260561713U002",
        WMS_HEADERS[2]: 1,
        WMS_HEADERS[3]: "1-1-01",
    },
]

# 2) La copia conserva las tres hojas, encabezados exactos y tipos de datos.
payload = export_official_putaway_xlsx(result.rows, TEMPLATE.read_bytes())
workbook = load_workbook(BytesIO(payload), data_only=False)
assert workbook.sheetnames == ["Sheet1", "Sheet2", "Sheet3"]
sheet = workbook["Sheet1"]
assert tuple(sheet.cell(1, column).value for column in range(1, 5)) == WMS_HEADERS
assert sheet["A2"].value == "PAS3902608080RT"
assert sheet["B2"].value == "MJ260510161U013"
assert sheet["C2"].value == 1 and isinstance(sheet["C2"].value, int)
assert sheet["D2"].value == "1-1-01"
for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=4):
    assert all(not (isinstance(cell.value, str) and cell.value.startswith("=")) for cell in row)

# 3) Una descarga incompleta requiere confirmación expresa.
partial_blocked = build_putaway_rows(
    events, records, "PAS3902608080RT", default_location="1-1-01",
    received=2, expected=949, allow_partial=False,
)
assert not partial_blocked.ready
assert any("carga parcial" in message.lower() for message in partial_blocked.errors)

# 4) Duplicados, ubicación faltante y fórmula potencial son bloqueantes.
duplicate = build_putaway_rows(
    events + [accepted("mj260561713u002", "MJ260561713")],
    records,
    "PAS3902608080RT",
    default_location="1-1-01",
)
assert not duplicate.ready and any("duplicado" in message.lower() for message in duplicate.errors)

missing_location = build_putaway_rows(events, records, "PAS3902608080RT")
assert not missing_location.ready and any("ubicación" in message.lower() for message in missing_location.errors)

formula_location = build_putaway_rows(
    events, records, "PAS3902608080RT", default_location="=1+1"
)
assert not formula_location.ready and any("comenzar" in message.lower() for message in formula_location.errors)

# 5) El código base repetible no se inventa como identificador individual.
base_event = [{
    "Escaneo": "MJ260510161",
    "Código": "MJ260510161",
    "Posición": "I01",
    "Tarima": "I01-01",
    "Estado": "OK",
    "Caja individual": False,
}]
base_result = build_putaway_rows(
    base_event, records, "PAS3902608080RT", default_location="1-1-01"
)
assert not base_result.ready and any("código base" in message.lower() for message in base_result.errors)

# 6) Una posición reutilizada puede dirigir dos tarimas a ubicaciones diferentes.
two_pallets = [
    accepted("MJ260510161U013", "MJ260510161", pallet="I01-01"),
    accepted("MJ260561713U002", "MJ260561713", pallet="I01-02"),
]
mapped = build_putaway_rows(
    two_pallets,
    records,
    "PAS3902608080RT",
    location_by_pallet={"I01-01": "1-1-01", "I01-02": "1-1-02"},
)
assert mapped.ready, mapped.errors
assert [row[WMS_HEADERS[3]] for row in mapped.rows] == ["1-1-01", "1-1-02"]

# 7) Normalización U2/U002, rango y persistencia del ID de tarima.
settings = Settings(target_capacity=1.94, physical_capacity=2.16)
live = LiveUnload([records["MJ260561713"]], settings, initial_left=1, initial_right=0)
scan = live.scan("mj260561713u2")
assert scan["ok"] and scan["scan"] == "MJ260561713U002"
assert live.history[0]["Tarima"] == "I01-01"
duplicate_scan = live.scan("MJ260561713U002")
assert not duplicate_scan["ok"] and duplicate_scan["status"] == "DUPLICADA"
out_of_range = live.scan("MJ260561713U007")
assert not out_of_range["ok"] and out_of_range["status"] == "FUERA DE RANGO"

restored = LiveUnload.from_state(
    [records["MJ260561713"]], settings, live.to_state()
)
assert restored.history == live.history
assert restored.scanned_unique_barcodes == live.scanned_unique_barcodes

print("OK WMS putaway: exportación oficial, validaciones y persistencia")

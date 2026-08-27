from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Iterable, Mapping
import re
import unicodedata

from openpyxl import load_workbook


WMS_HEADERS = (
    "Putaway Order/上架单号",
    "Box Type or Custom Box Barcode/箱类型号or自定义箱条码",
    "Putaway Qty/上架数",
    "Location/上架库位",
)

REJECTED_STATES = {
    "DUPLICADA",
    "NO ENCONTRADA",
    "SOBRANTE",
    "SIN POSICIÓN",
    "POSICIÓN PENDIENTE",
    "VACÍO",
    "FUERA DE RANGO",
}

_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MAX_EXCEL_DATA_ROWS = 1_048_575


@dataclass
class WmsBuildResult:
    rows: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def ready(self) -> bool:
        return bool(self.rows) and not self.errors


def _normalized_text(value) -> str:
    text = "" if value is None else str(value)
    return unicodedata.normalize("NFKC", text).strip()


def _canonical_identifier(value) -> str:
    return re.sub(r"\s+", "", _normalized_text(value)).upper()


def _validate_text(value, field_name: str, errors: list[str]) -> str:
    text = _normalized_text(value)
    if not text:
        errors.append(f"{field_name}: el valor está vacío.")
        return ""
    if "\n" in text or "\r" in text or _CONTROL_CHARS.search(text):
        errors.append(f"{field_name}: contiene saltos de línea o caracteres de control.")
    if text.startswith(_FORMULA_PREFIXES):
        errors.append(f"{field_name}: no puede comenzar con =, +, - o @.")
    return text


def pallet_id_for_event(event: Mapping) -> str:
    pallet_id = _normalized_text(event.get("Tarima", ""))
    if pallet_id:
        return pallet_id
    position = _normalized_text(event.get("Posición", ""))
    return f"{position}-01" if position else ""


def summarize_pallets(accepted_events: Iterable[Mapping]) -> list[dict]:
    """Resume únicamente eventos aceptados, en el orden de la primera caja."""
    grouped: dict[str, dict] = {}
    for event in accepted_events:
        if _normalized_text(event.get("Estado", "")).upper() in REJECTED_STATES:
            continue
        pallet_id = pallet_id_for_event(event)
        if not pallet_id:
            continue
        if pallet_id not in grouped:
            grouped[pallet_id] = {
                "Tarima": pallet_id,
                "Posición física": _normalized_text(event.get("Posición", "")),
                "Cajas": 0,
                "Primera caja": _canonical_identifier(event.get("Escaneo", "")),
            }
        grouped[pallet_id]["Cajas"] += 1
    return list(grouped.values())


def validate_putaway_rows(rows: Iterable[Mapping]) -> list[str]:
    errors: list[str] = []
    seen: set[str] = set()
    rows = list(rows)
    if not rows:
        return ["No hay cajas para incluir en la plantilla."]
    if len(rows) > _MAX_EXCEL_DATA_ROWS:
        errors.append("La cantidad de filas excede el límite de una hoja de Excel.")

    for index, row in enumerate(rows, start=2):
        order = _validate_text(row.get(WMS_HEADERS[0]), f"Fila {index}, orden Putaway", errors)
        barcode = _validate_text(row.get(WMS_HEADERS[1]), f"Fila {index}, código de caja", errors)
        location = _validate_text(row.get(WMS_HEADERS[3]), f"Fila {index}, ubicación WMS", errors)
        qty = row.get(WMS_HEADERS[2])
        if qty != 1 or isinstance(qty, bool):
            errors.append(f"Fila {index}, cantidad: debe ser el número entero 1.")
        canonical = _canonical_identifier(barcode)
        if canonical:
            if canonical in seen:
                errors.append(f"Fila {index}: el código de caja {barcode} está duplicado.")
            seen.add(canonical)
        if order and _canonical_identifier(order) != order.upper():
            errors.append(f"Fila {index}, orden Putaway: no debe contener espacios.")
        if location and location != location.strip():
            errors.append(f"Fila {index}, ubicación WMS: contiene espacios exteriores.")
    return errors


def build_putaway_rows(
    accepted_events: Iterable[Mapping],
    records_by_code: Mapping[str, object],
    putaway_order: str,
    *,
    default_location: str = "",
    location_by_pallet: Mapping[str, str] | None = None,
    location_by_position: Mapping[str, str] | None = None,
    received: int | None = None,
    expected: int | None = None,
    allow_partial: bool = False,
    require_final_validation: bool = True,
) -> WmsBuildResult:
    """Construye filas estrictas de caja individual para la plantilla oficial.

    La ubicación se resuelve por tarima, después por posición física y finalmente
    con la ubicación predeterminada. Nunca se interpreta I01/D01 como ubicación WMS.
    """
    result = WmsBuildResult()
    order = _validate_text(putaway_order, "Orden Putaway", result.errors)
    if order:
        order = _canonical_identifier(order)
        if not order.startswith("PAS"):
            result.warnings.append(
                "La orden no comienza con PAS. Verifique que sea la orden Putaway mostrada por el WMS."
            )

    default_location = _normalized_text(default_location)
    location_by_pallet = location_by_pallet or {}
    location_by_position = location_by_position or {}

    if received is not None and expected is not None and received < expected:
        result.warnings.append(
            f"Carga parcial: se incluirán {received} de {expected} cajas esperadas."
        )
        if not allow_partial:
            result.errors.append(
                "La descarga está incompleta. Active la confirmación de carga parcial para continuar."
            )

    accepted = [
        event for event in accepted_events
        if _normalized_text(event.get("Estado", "")).upper() not in REJECTED_STATES
    ]
    if not accepted:
        result.errors.append("No hay cajas aceptadas para generar la plantilla.")
        return result

    seen: set[str] = set()
    for sequence, event in enumerate(accepted, start=1):
        barcode = _canonical_identifier(event.get("Escaneo", ""))
        base_code = _canonical_identifier(event.get("Código", ""))
        pallet_id = pallet_id_for_event(event)
        physical_position = _normalized_text(event.get("Posición", ""))

        if require_final_validation and event.get("Elegible WMS") is not True:
            physical_state = _normalized_text(event.get("Estado físico", "SIN CONFIRMACIÓN"))
            validated = event.get("Tarima validada") is True
            result.errors.append(
                f"Caja {sequence}: {barcode or 'sin código'} no es elegible para WMS "
                f"(estado {physical_state or 'SIN CONFIRMACIÓN'}, "
                f"tarima {'validada' if validated else 'no validada'})."
            )
            continue

        if not barcode:
            result.errors.append(f"Caja {sequence}: el escaneo está vacío.")
            continue
        _validate_text(barcode, f"Caja {sequence}, código", result.errors)
        if barcode in seen:
            result.errors.append(f"Caja {sequence}: el código {barcode} está duplicado.")
            continue
        seen.add(barcode)

        record = records_by_code.get(base_code)
        expected_boxes = getattr(record, "boxes", None)
        unique_flag = event.get("Caja individual")
        if unique_flag in ("", None):
            unique_flag = barcode != base_code or expected_boxes == 1
        if not bool(unique_flag) and (expected_boxes or 0) > 1:
            result.errors.append(
                f"Caja {sequence}: {barcode} es el código base de un grupo de "
                f"{expected_boxes} cajas; se requiere el código individual real."
            )
            continue

        location = _normalized_text(location_by_pallet.get(pallet_id, ""))
        if not location:
            location = _normalized_text(location_by_position.get(physical_position, ""))
        if not location:
            location = default_location
        if not location:
            result.errors.append(
                f"Tarima {pallet_id or 'sin identificar'}: falta la ubicación WMS."
            )
            continue
        _validate_text(location, f"Tarima {pallet_id}, ubicación WMS", result.errors)

        result.rows.append({
            WMS_HEADERS[0]: order,
            WMS_HEADERS[1]: barcode,
            WMS_HEADERS[2]: 1,
            WMS_HEADERS[3]: location,
        })

    result.errors.extend(validate_putaway_rows(result.rows))
    result.errors = list(dict.fromkeys(result.errors))
    result.warnings = list(dict.fromkeys(result.warnings))
    return result


def export_official_putaway_xlsx(rows: Iterable[Mapping], template_bytes: bytes) -> bytes:
    """Rellena una copia en memoria de la plantilla oficial sin cambiar su diseño."""
    rows = list(rows)
    errors = validate_putaway_rows(rows)
    if errors:
        raise ValueError("No se puede exportar: " + " | ".join(errors))

    workbook = load_workbook(BytesIO(template_bytes))
    if "Sheet1" not in workbook.sheetnames:
        raise ValueError("La plantilla oficial no contiene la hoja Sheet1.")
    sheet = workbook["Sheet1"]
    headers = tuple(sheet.cell(1, column).value for column in range(1, 5))
    if headers != WMS_HEADERS:
        raise ValueError("Los encabezados de la plantilla no coinciden con el formato oficial esperado.")

    if sheet.max_row > 1:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.value = None

    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row_index, 1, str(row[WMS_HEADERS[0]])).number_format = "@"
        sheet.cell(row_index, 2, str(row[WMS_HEADERS[1]])).number_format = "@"
        sheet.cell(row_index, 3, 1).number_format = "0"
        sheet.cell(row_index, 4, str(row[WMS_HEADERS[3]])).number_format = "@"

    final_row = len(rows) + 1
    for table in sheet.tables.values():
        table.ref = f"A1:D{final_row}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

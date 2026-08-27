from io import BytesIO
import json

from openpyxl import load_workbook

from core.live import LiveUnload
from core.optimizer import Settings, transfer_layout_summary
from core.parser import CodeRecord
from core.pda_exchange import demo_pda_result, parse_pda_result
from core.wms_putaway import WMS_HEADERS, build_putaway_rows, export_official_putaway_xlsx


records = {
    "MJ260510161": CodeRecord("MJ260510161", 3, 0.30, 0.10),
    "UNITARIO": CodeRecord("UNITARIO", 1, 0.05, 0.05),
}
settings = Settings()


# El código base de un grupo nunca se contabiliza.
live = LiveUnload(list(records.values()), settings, initial_left=1, initial_right=1)
base = live.scan("MJ260510161")
assert not base["ok"] and base["status"] == "LECTURA INCOMPLETA"
assert live.received["MJ260510161"] == 0

# U2 se normaliza a U002; una lectura corrupta del mismo barcode es duplicada.
accepted = live.scan("MJ260510161U2")
assert accepted["ok"] and accepted["scan"] == "MJ260510161U002"
duplicate = live.scan("XXXMJ260510161U002MJ260510161U002")
assert not duplicate["ok"] and duplicate["status"] == "DUPLICADA"
assert live.received["MJ260510161"] == 1

# El código base unitario sí es inequívoco y se convierte a U001.
unit = live.scan("UNITARIO")
assert unit["ok"] and unit["scan"] == "UNITARIOU001"


# Resultado PDA válido: firma, contenedor, barcode, rango y tarima final.
pda_bytes = demo_pda_result(
    "CXDU2223616",
    records,
    [
        {
            "raw_scan": "MJ260510161U002",
            "barcode": "MJ260510161U002",
            "code": "MJ260510161",
            "box_number": 2,
            "final_pallet": "T-01",
            "transfer_pallet": "TR-01",
            "direct_to_final": False,
            "scanned_at": "2026-08-27 10:00:00",
        },
        {
            "raw_scan": "UNITARIO",
            "barcode": "UNITARIOU001",
            "code": "UNITARIO",
            "box_number": 1,
            "final_pallet": "T-02",
            "transfer_pallet": "",
            "direct_to_final": True,
            "scanned_at": "2026-08-27 10:01:00",
        },
    ],
)
imported = parse_pda_result(pda_bytes, records, "CXDU2223616")
assert imported.ready, imported.errors
assert imported.schema_version == 2 and len(imported.eligible_events) == 2
assert [row["Tarima"] for row in imported.events] == ["T-01", "T-02"]
assert [row["Escaneo"] for row in imported.events] == ["MJ260510161U002", "UNITARIOU001"]


# El resultado validado alimenta directamente el generador oficial WMS.
wms = build_putaway_rows(
    imported.events,
    records,
    "PAS3902608080RT",
    location_by_pallet={"T-01": "2B-F03-02", "T-02": "2B-F03-03"},
    received=2,
    expected=4,
    allow_partial=True,
)
assert wms.ready, wms.errors
assert [row[WMS_HEADERS[3]] for row in wms.rows] == ["2B-F03-02", "2B-F03-03"]

from pathlib import Path

template = Path("assets/templates/Plantilla_oficial_WMS_PutawayCrossDockImport.xlsx").read_bytes()
payload = export_official_putaway_xlsx(wms.rows, template)
workbook = load_workbook(BytesIO(payload), data_only=False)
assert tuple(workbook["Sheet1"].cell(1, c).value for c in range(1, 5)) == WMS_HEADERS
assert workbook["Sheet1"]["B2"].value == "MJ260510161U002"
assert workbook["Sheet1"]["D3"].value == "2B-F03-03"


# La preparación física separa tendido, directas al pie y TR-01.
layout_records = [
    CodeRecord("GRANDEA", 5, 2.50, 0.50),
    CodeRecord("GRANDEB", 2, 1.50, 0.75),
    CodeRecord("CHICO", 3, 0.30, 0.10),
]
layout = transfer_layout_summary(layout_records, settings, foot_positions=1)
assert layout["tendido_final"] == 1
assert layout["foot_direct_initial"] == 1 and layout["foot_transfer_initial"] == 1
assert layout["physical_initial_total"] == 3
assert layout["direct_replacements"] >= 1


# Una caja todavía en TR-xx se importa para auditoría, pero queda bloqueada para WMS.
transit_bytes = demo_pda_result(
    "CXDU2223616",
    records,
    [{
        "raw_scan": "MJ260510161U001",
        "barcode": "MJ260510161U001",
        "code": "MJ260510161",
        "box_number": 1,
        "final_pallet": "T-01",
        "transfer_pallet": "TR-01",
        "direct_to_final": False,
        "physical_state": "EN_TRASLADO",
        "transfer_distributed": False,
        "final_pallet_validated": False,
        "wms_eligible": False,
    }],
)
transit = parse_pda_result(transit_bytes, records, "CXDU2223616")
assert transit.ready and len(transit.eligible_events) == 0
blocked_wms = build_putaway_rows(
    transit.events, records, "PAS3902608080RT", default_location="2B-F03-02",
    received=1, expected=4, allow_partial=True,
)
assert not blocked_wms.ready and any("no es elegible" in error.lower() for error in blocked_wms.errors)


# Otro contenedor o Packing List quedan bloqueados.
wrong_container = parse_pda_result(pda_bytes, records, "OTRO1234567")
assert not wrong_container.ready and any("corresponde" in error for error in wrong_container.errors)

other_records = dict(records)
other_records["MJ260510161"] = CodeRecord("MJ260510161", 4, 0.40, 0.10)
wrong_manifest = parse_pda_result(pda_bytes, other_records, "CXDU2223616")
assert not wrong_manifest.ready and any("otro Packing List" in error for error in wrong_manifest.errors)

# Un campo manipulado no debe provocar una excepción ni llegar al WMS.
tampered_payload = json.loads(pda_bytes)
tampered_payload["accepted_events"][0]["box_number"] = "NO-ES-NÚMERO"
tampered = parse_pda_result(
    json.dumps(tampered_payload).encode("utf-8"), records, "CXDU2223616"
)
assert not tampered.ready and any("declarado es inválido" in error for error in tampered.errors)

print("OK V0.9 Windows: U001-UN + estado físico + archivo WMS exclusivo")

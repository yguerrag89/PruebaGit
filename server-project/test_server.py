"""Synthetic integration tests: auth, native v4, offline retries, export and backup."""
import concurrent.futures
import copy
import io
import json
import re
import sqlite3
import tempfile
import unittest
import uuid
from pathlib import Path
from fastapi.testclient import TestClient
from openpyxl import load_workbook, Workbook

from app import create_app
from store import Store, encode
from protocol import create_session, claim_session, accept_snapshot, export_wms, Rejected
from core.parser import CodeRecord, ParsedContainer
from core.optimizer import Settings
from core.pda_exchange import build_pda_manifest
from core.wms_putaway import WMS_HEADERS

FIXTURES = Path(__file__).resolve().parent.parent / "windows-project/tests/fixtures/v011"


class ServerTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.path = Path(self.temp.name) / "data.sqlite3"
        self.store = Store(self.path)
        self.store.initialize("Synthetic-test-pass-2026")
        self.records = [CodeRecord("CHICOA",3,.6,.2),CodeRecord("CHICOB",3,.6,.2)]
        self.manifest = json.loads(build_pda_manifest(ParsedContainer("CONTINUO","test.xlsx","",self.records,[]),Settings()))
        self.sid, pairing = create_session(self.store,self.manifest,1,0)
        self.token = pairing.split(".")[1]
        self.device = str(uuid.uuid4())
        self.claim = claim_session(self.store,self.sid,self.token,self.device)
        self.client = TestClient(create_app(self.path,"https://testserver"),base_url="https://testserver")

    def tearDown(self):
        self.client.close()
        self.temp.cleanup()

    def packet(self, name="v4-verified.json", revision=1, sealed=False):
        result = json.loads((FIXTURES/name).read_bytes())
        return dict(schema="ilubox.sync.v1",session_id=self.sid,manifest_hash=self.claim["manifest_hash"],
                    revision=revision,sealed=sealed,partial_reason="",result=result,
                    audit=[dict(id=i,barcode=e["barcode"],accepted=True,scan=e["raw_scan"]) for i,e in enumerate(result["accepted_events"],1)])

    def send(self, packet):
        return accept_snapshot(self.store,self.sid,self.token,self.device,encode(packet))

    def login(self):
        response = self.client.post("/login",data={"password":"Synthetic-test-pass-2026"},headers={"Origin":"https://testserver"})
        self.assertEqual(200,response.status_code,response.text)
        return re.search('name="csrf" value="([^"]+)"',response.text).group(1)

    def test_retry_and_final_export(self):
        packet = self.packet(sealed=True)
        first = self.send(packet)
        # Simulate losing the HTTP response and restarting the server process.
        self.store = Store(self.path)
        self.assertEqual(first,self.send(packet))
        self.assertEqual(6,json.loads(self.store.get(self.sid)["payload"])["result"]["progress"]["received"])
        with self.assertRaises(Rejected): self.send(self.packet(revision=2))
        content, hash_value = export_wms(self.store,self.sid,"PAS-TEST")
        self.assertEqual((content,hash_value),export_wms(self.store,self.sid,"PAS-TEST"))
        with self.assertRaises(Rejected): export_wms(self.store,self.sid,"PAS-OTRA")
        sheet = load_workbook(io.BytesIO(content)).active
        self.assertEqual(WMS_HEADERS,tuple(c.value for c in sheet[1]))
        self.assertEqual(7,sheet.max_row)
        self.assertEqual({"2B-TMP-01","2B-TMP-02"},{sheet.cell(i,4).value for i in range(2,8)})
        self.assertTrue(all(sheet.cell(i,3).value==1 for i in range(2,8)))

    def test_old_revisions_and_conflicts(self):
        self.send(self.packet(revision=2))
        with self.assertRaises(Rejected): self.send(self.packet(revision=1))
        altered=self.packet(revision=2); altered["partial_reason"]="different"
        with self.assertRaises(Rejected): self.send(altered)
        self.assertEqual(2,self.store.get(self.sid)["revision"])

    def test_other_device_and_list(self):
        with self.assertRaises(Rejected): claim_session(self.store,self.sid,self.token,str(uuid.uuid4()))
        with self.assertRaises(Rejected): accept_snapshot(self.store,self.sid,"bad",self.device,encode(self.packet()))
        altered=self.packet(); altered["manifest_hash"]="wrong"
        with self.assertRaises(Rejected): self.send(altered)
        with self.assertRaises(sqlite3.IntegrityError): create_session(self.store,self.manifest,1,0)
        self.assertEqual(self.claim,claim_session(self.store,self.sid,self.token,self.device))

    def test_pending_and_partial(self):
        packet=self.packet("v4-pending.json")
        self.send(packet)
        with self.assertRaises(Rejected): export_wms(self.store,self.sid,"PAS-TEST")
        packet["sealed"]=True;packet["revision"]=2
        with self.assertRaises(Rejected): self.send(packet)

    def test_immutable_tarima_and_audit(self):
        self.send(self.packet())
        changed=self.packet(revision=2);changed["audit"]=[]
        with self.assertRaises(Rejected): self.send(changed)
        changed=self.packet(revision=2)
        for p in changed["result"]["pallets"]: p["wms_temporary_location"]="2B-CAMBIO"
        for e in changed["result"]["accepted_events"]: e["wms_temporary_location"]="2B-CAMBIO"
        with self.assertRaises(Rejected): self.send(changed)
        changed=self.packet(revision=3);changed["result"]["accepted_events"].append(copy.deepcopy(changed["result"]["accepted_events"][0]))
        with self.assertRaises(Rejected): self.send(changed)

    def test_parallel_retry_one_snapshot(self):
        packet=self.packet()
        with concurrent.futures.ThreadPoolExecutor(4) as pool:
            results=list(pool.map(lambda _: self.send(packet),range(12)))
        self.assertTrue(all(x==results[0] for x in results))
        self.assertEqual(1,self.store.get(self.sid)["revision"])

    def test_web_auth_csrf_and_download(self):
        self.assertEqual(401,self.client.get(f"/sessions/{self.sid}").status_code)
        self.assertEqual(403,self.client.post("/login",data={"password":"Synthetic-test-pass-2026"}).status_code)
        csrf=self.login()
        detail=self.client.get(f"/sessions/{self.sid}")
        self.assertIn("Preparación del espacio",detail.text)
        self.assertEqual(403,self.client.post(f"/sessions/{self.sid}/export",headers={"Origin":"https://testserver"},data={"csrf":"bad"}).status_code)
        self.send(self.packet(sealed=True))
        response=self.client.post(f"/sessions/{self.sid}/export",headers={"Origin":"https://testserver"},data={"csrf":csrf,"order":"PAS-TEST","reviewed":"yes"})
        self.assertEqual(200,response.status_code,response.text[:100])
        self.assertIn("X-Content-SHA256",response.headers)
        self.assertEqual(400,self.client.get("/",headers={"host":"public.invalid"}).status_code)
        self.assertEqual(413,self.client.put(f"/api/sessions/{self.sid}/snapshot",headers={"content-length":"20000000"}).status_code)
        self.assertEqual("no-store",detail.headers["Cache-Control"])

    def test_api_and_backup(self):
        headers={"Authorization":"Bearer "+self.token,"X-Ilubox-Device":self.device}
        response=self.client.put(f"/api/sessions/{self.sid}/snapshot",headers=headers,content=encode(self.packet()))
        self.assertEqual(200,response.status_code,response.text)
        self.assertEqual(401,self.client.put(f"/api/sessions/{self.sid}/snapshot",content=b"{}").status_code)
        backup=Path(self.temp.name)/"backup.sqlite3"
        self.store.backup(backup)
        self.assertEqual(self.store.get(self.sid),Store(backup).get(self.sid))
        with self.assertRaises(FileExistsError): self.store.backup(backup)

    def test_malformed_json_and_xlsx(self):
        for content in (b'[]',b'null',b'{',b'{"revision":true}'):
            with self.assertRaises(Rejected): accept_snapshot(self.store,self.sid,self.token,self.device,content)
        csrf=self.login()
        response=self.client.post("/sessions",headers={"Origin":"https://testserver"},data={"csrf":csrf},files={"packing":("bad.xlsx",b"notxlsx")})
        self.assertEqual(422,response.status_code)

    def test_upload_and_reject_fractional_counts(self):
        csrf=self.login()
        for count,expected in [(1.5,422),(2,200)]:
            book=Workbook();sheet=book.active
            sheet.append(["Codigo","Cajas","CBM","Contenedor"])
            sheet.append(["NUEVA",count,.2,"MSKU1234567"])
            content=io.BytesIO();book.save(content)
            response=self.client.post("/sessions",headers={"Origin":"https://testserver"},data={"csrf":csrf,"left":1,"right":0},files={"packing":("packing.xlsx",content.getvalue())})
            self.assertEqual(expected,response.status_code,response.text[:500])

    def test_unclaimed_assignment_correction(self):
        csrf=self.login()
        manifest=copy.deepcopy(self.manifest);manifest["container_id"]="SIN-ASIGNAR"
        sid,pairing=create_session(self.store,manifest,1,0)
        headers={"Origin":"https://testserver"}
        response=self.client.post(f"/sessions/{sid}/assignment",headers=headers,data={"csrf":csrf,"action":"reissue"})
        self.assertEqual(200,response.status_code)
        with self.assertRaises(Rejected): claim_session(self.store,sid,pairing.split(".")[1],str(uuid.uuid4()))
        response=self.client.post(f"/sessions/{sid}/assignment",headers=headers,data={"csrf":csrf,"action":"cancel"})
        self.assertEqual(200,response.status_code)
        with self.assertRaises(Rejected): export_wms(self.store,sid,"PAS-TEST")
        create_session(self.store,manifest,1,0)
        self.assertEqual(409,self.client.post(f"/sessions/{self.sid}/assignment",headers=headers,data={"csrf":csrf,"action":"cancel"}).status_code)


if __name__=="__main__":
    unittest.main()

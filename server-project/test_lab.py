"""End-to-end checks for the isolated Windows laboratory."""
import io
import json
import re
import tempfile
import unittest
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from lab import create_lab_app
from lab_engine import initial_state, records, release_pallet, scan, snapshot_result, validate_pallet
from core.optimizer import Settings
from core.parser import CodeRecord, ParsedContainer
from core.pda_exchange import build_pda_manifest, parse_pda_result
from core.wms_putaway import WMS_HEADERS


ORIGIN = "http://127.0.0.1:8765"


def packing_bytes(rows):
    book = Workbook()
    sheet = book.active
    sheet.title = "Packing"
    sheet.append(["Codigo", "Cajas", "CBM", "Contenedor"])
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    book.save(output)
    return output.getvalue()


class LabWebTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.client = TestClient(create_lab_app(self.root, ORIGIN), base_url=ORIGIN)
        response = self.client.get("/")
        self.assertEqual(200, response.status_code)
        self.csrf = re.search(r'name="csrf" value="([^"]+)"', response.text).group(1)

    def tearDown(self):
        self.client.close()
        self.temp.cleanup()

    def post(self, path, data=None, files=None):
        payload = {"csrf": self.csrf}
        payload.update(data or {})
        return self.client.post(path, data=payload, files=files,
                                headers={"Origin": ORIGIN}, follow_redirects=True)

    def state(self):
        return json.loads((self.root / "pda-simulada.json").read_text(encoding="utf-8"))

    def meta(self):
        return json.loads((self.root / "laboratorio.json").read_text(encoding="utf-8"))

    def start_three_boxes(self):
        content = packing_bytes([
            ["CHICOA", 2, 0.4, "MSKU1234567"],
            ["CHICOB", 1, 0.2, "MSKU1234567"],
        ])
        response = self.post("/start", {"left": "1", "right": "1"},
                             {"packing": ("MSKU1234567.xlsx", content,
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        self.assertEqual(200, response.status_code, response.text[:500])
        self.assertIn("MSKU1234567", response.text)

    def test_complete_operational_flow_offline_retry_and_wms(self):
        self.start_three_boxes()
        initial = self.state()
        self.assertEqual(0, initial["revision"])
        self.assertEqual({"T-01"}, set(initial["planned"].values()))

        response = self.post("/scan", {"scan": "CHICOAU000"})
        self.assertIn("El número de caja debe ser mayor que cero", response.text)
        self.assertEqual(0, len(self.state()["boxes"]))
        self.assertEqual(self.state()["revision"], self.state()["acknowledged"])

        self.post("/scan", {"scan": "CHICOAU001"})
        self.post("/scan", {"scan": "CHICOAU001"})
        state = self.state()
        self.assertEqual(1, len(state["boxes"]))
        self.assertIn("YA ESCANEADA", state["last_message"])

        self.post("/network")
        before_offline = self.state()["acknowledged"]
        self.post("/scan", {"scan": "CHICOAU002"})
        self.post("/scan", {"scan": "CHICOBU001"})
        offline = self.state()
        self.assertFalse(offline["online"])
        self.assertGreater(offline["revision"], offline["acknowledged"])
        self.assertEqual(before_offline, offline["acknowledged"])
        self.assertIn("pending", self.meta())

        self.post("/network")
        online = self.state()
        self.assertTrue(online["online"])
        self.assertEqual(online["revision"], online["acknowledged"])
        self.assertNotIn("pending", self.meta())

        self.post("/transfer")
        incomplete = self.post("/validate", {"pallet": "T-01", "responsible": "Operador Uno",
                                               "temporary": "I01"})
        self.assertIn("Temporal WMS inválida", incomplete.text)
        response = self.post("/validate", {"pallet": "t-01", "responsible": "Operador Uno",
                                            "temporary": "2b-tmp-01"})
        self.assertIn("T-01 verificada", response.text)
        self.assertEqual(3, snapshot_result(self.state())["progress"]["wms_eligible"])

        retry = self.post("/sync", {"twice": "yes"})
        self.assertIn("Reintento idempotente aprobado", retry.text)
        self.assertEqual(3, len(self.state()["boxes"]))

        closed = self.post("/close", {"reason": ""})
        self.assertIn("Descarga cerrada", closed.text)
        self.assertTrue(self.state()["sealed"])
        self.assertEqual(self.state()["revision"], self.state()["acknowledged"])

        exported = self.post("/export", {"order": "PAS-LAB-001"})
        self.assertEqual(200, exported.status_code, exported.text[:200])
        self.assertIn("X-Content-SHA256", exported.headers)
        sheet = load_workbook(io.BytesIO(exported.content)).active
        self.assertEqual(WMS_HEADERS, tuple(cell.value for cell in sheet[1]))
        self.assertEqual(4, sheet.max_row)
        self.assertEqual({"CHICOAU001", "CHICOAU002", "CHICOBU001"},
                         {sheet.cell(row, 2).value for row in range(2, 5)})
        self.assertTrue(all(sheet.cell(row, 3).value == 1 for row in range(2, 5)))
        self.assertTrue(all(sheet.cell(row, 4).value == "2B-TMP-01" for row in range(2, 5)))

    def test_strict_upload_persists_after_app_restart_and_reset(self):
        bad = packing_bytes([["MALO", 1.5, 0.2, "MSKU1234567"]])
        response = self.post("/start", {"left": "1", "right": "0"},
                             {"packing": ("bad.xlsx", bad,
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        self.assertEqual(422, response.status_code)
        self.assertIn("no se omitió la fila", response.text)

        self.start_three_boxes()
        self.post("/scan", {"scan": "CHICOAU001"})
        self.client.close()
        self.client = TestClient(create_lab_app(self.root, ORIGIN), base_url=ORIGIN)
        response = self.client.get("/")
        self.assertIn("1/3", response.text)
        self.csrf = re.search(r'name="csrf" value="([^"]+)"', response.text).group(1)
        self.post("/reset", {"confirm": "BORRAR PRUEBA"})
        self.assertFalse((self.root / "pda-simulada.json").exists())
        self.assertIn("Crear prueba con un Packing List", self.client.get("/").text)

    def test_packing_upload_does_not_depend_on_browser_origin_header(self):
        content = packing_bytes([["CHICO", 1, 0.2, "MSKU1234567"]])
        response = self.client.post(
            "/start",
            data={"csrf": self.csrf, "left": "1", "right": "0"},
            files={"packing": ("MSKU1234567.xlsx", content,
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=True,
        )
        self.assertEqual(200, response.status_code, response.text[:500])
        self.assertIn("MSKU1234567", response.text)

        expired = self.client.post("/network", data={"csrf": "formulario-vencido"})
        self.assertEqual(403, expired.status_code)
        self.assertIn("Ya se generó un formulario nuevo", expired.text)
        self.assertNotIn('{"detail":', expired.text)


class LabEngineDirectTest(unittest.TestCase):
    def test_direct_pallet_must_be_released_before_reusing_position(self):
        manifest = json.loads(build_pda_manifest(
            ParsedContainer("MSKU7654321", "synthetic.xlsx", "Packing",
                            [CodeRecord("GRANDE", 4, 2.0, 0.5)], []), Settings()))
        state = initial_state(manifest, 1, 0)
        self.assertEqual(["GRANDE"], state["direct_codes"])
        for number in range(1, 4):
            self.assertTrue(scan(state, f"GRANDEU{number:03d}"))
        self.assertFalse(scan(state, "GRANDEU004"))
        self.assertEqual(3, len(state["boxes"]))
        self.assertIn("Verifique y retire T-01", state["last_message"])

        validate_pallet(state, "T-01", "Operador", "TMP/PIE/01")
        release_pallet(state, "T-01")
        self.assertTrue(scan(state, "GRANDEU004"))
        validate_pallet(state, "T-02", "Operador", "TMP/PIE/02")
        result = snapshot_result(state)
        parsed = parse_pda_result(json.dumps(result).encode(), records(state), "MSKU7654321")
        self.assertFalse(parsed.errors, parsed.errors)
        self.assertEqual(4, len(parsed.eligible_events))
        self.assertEqual("I01", next(p for p in result["pallets"] if p["id"] == "T-02")["physical_position"])


if __name__ == "__main__":
    unittest.main()
